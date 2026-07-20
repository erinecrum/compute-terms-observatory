"""Export the comparison dataset to a downloadable Excel workbook.

Two sheets:
  * "Comparison" - the matrix grid (section, dimension, one column per provider),
    the same view as the site, for a quick side-by-side read.
  * "Detail" - one row per provider x dimension with the full provenance an
    attorney needs to verify: value, confidence, whether it was human-verified,
    the citation, the source document name, its URL, the fetch date, and the
    content-hash of the archived version.

PDF is produced from the site itself via the browser's Print / Save as PDF (the
print stylesheet formats the page for paper), so there is no PDF code here.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import is_applicable
from .site import _visible_dims  # Phase C render-gate (hides new groups until Phase D)

_INK = "16202E"
_ACCENT = "1C3F63"
_PANEL = "EAEFF7"
_LINE = "D5DCE7"

_HEADER_FILL = PatternFill("solid", fgColor=_ACCENT)
_SECTION_FILL = PatternFill("solid", fgColor=_PANEL)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(bold=True, color=_ACCENT, size=10)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color=_LINE)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

DISCLAIMER_SHORT = (
    "AI-generated summaries of public documents, not legal advice. Values can be "
    "wrong or out of date. Verify each against its linked source before relying on it."
)

# Shown whenever a sheet carries rows from the data-protection or accountability
# groups. These dimensions describe the mechanism a provider has published; they
# are not, and must not be read as, an assessment of compliance with any law.
NEW_GROUP_NOTE = (
    "Note on the 'Data protection & privacy' and 'Accountability & transparency' "
    "rows: these are descriptive extractions of the mechanisms a provider has "
    "published, not compliance assessments. Nothing here states or implies that a "
    "provider does or does not comply with the GDPR, the EU AI Act, or any other law."
)
_NEW_GROUPS = {"Data protection & privacy", "Accountability & transparency"}

# Readable labels for the derived four-state status (Issue 2). Everything is
# AI-reviewed; there is no human-verified tier.
_STATUS_LABEL = {
    "quote_verified": "quote verified",
    "quote_unverified": "unverified (no quote matched)",
    "no_clause_found": "silent",
    "not_applicable": "not applicable",
}


def _date(s: str) -> str:
    return (s or "")[:10]


def _comparison_sheet(ws, dataset: dict, group: str, title: str, view: str = "all") -> None:
    from .schema import is_applicable
    from .site import view_dims

    # `group` here is the display SECTION. Providers are selected by section, but
    # the dimension set follows the section's governing instrument, so Hosted
    # platforms takes the hosted-platform term set while sitting under Open.
    from .schema import SECTION_DIM_GROUP
    dim_group = SECTION_DIM_GROUP.get(group, group)
    providers = [p for p in dataset["providers"] if p.get("section") == group]
    dims = [d for d in _visible_dims(dataset["dimensions"]) if is_applicable(dim_group, d["key"])]
    dims = view_dims(dims, view)
    matrix = dataset["matrix"]

    ws.title = title
    ws["A1"] = f"Compute Terms Observatory — {title}"
    ws["A1"].font = Font(bold=True, size=14, color=_INK)
    ws["A2"] = DISCLAIMER_SHORT
    ws["A2"].font = Font(italic=True, size=9, color="586377")
    ws["A3"] = f"Data current as of {_date(dataset.get('data_current_as_of',''))}"
    ws["A3"].font = Font(size=9, color="586377")

    if not providers:
        return
    # The compliance-assessment note, only on sheets that actually carry those rows.
    header_row = 5
    if any(d.get("group") in _NEW_GROUPS for d in dims):
        ws["A4"] = NEW_GROUP_NOTE
        ws["A4"].font = Font(italic=True, size=9, color="586377")
        header_row = 6
    headers = ["Section", "Term dimension"] + [p["provider_name"] for p in providers]
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP_TOP
        cell.border = _BORDER

    r = header_row + 1
    for d in dims:
        ws.cell(row=r, column=1, value=d.get("group", "")).font = _SECTION_FONT
        ws.cell(row=r, column=1).fill = _SECTION_FILL
        ws.cell(row=r, column=2, value=d["label"]).font = Font(bold=True, size=10)
        for c, p in enumerate(providers, start=3):
            f = matrix.get(p["provider"], {}).get(d["key"], {})
            # A spreadsheet has no drawer to open, so an absence carries its full
            # sentence rather than the shortened matrix clause.
            ws.cell(row=r, column=c,
                    value=f.get("absence_full")
                    or f.get("display_value", f.get("value", "")))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 40


def _detail_sheet(ws, dataset: dict) -> None:
    providers = dataset["providers"]
    dims = _visible_dims(dataset["dimensions"])
    matrix = dataset["matrix"]

    ws.title = "Detail"
    note_rows = 0
    if any(d.get("group") in _NEW_GROUPS for d in dims):
        ws["A1"] = NEW_GROUP_NOTE
        ws["A1"].font = Font(italic=True, size=9, color="586377")
        note_rows = 2  # the note, then a blank spacer row
    headers = [
        "Provider", "Section", "Term dimension", "Value", "Status", "Confidence",
        "Citation", "Source document", "Source URL",
        "Fetched", "Version hash",
    ]
    header_row = 1 + note_rows
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP_TOP
        cell.border = _BORDER

    r = header_row + 1
    for p in providers:
        for d in dims:
            if not is_applicable(p.get("group", "cloud"), d["key"]):
                continue  # dimension structurally inapplicable to this segment
            f = matrix.get(p["provider"], {}).get(d["key"], {})
            src = f.get("source") or {}
            prog = f.get("commitment_program")
            value = (f.get("absence_full")
                     or f.get("display_value", f.get("value", "")))
            if prog:
                value = f"{value}\n[{prog['program']}: {prog['value']}]"
            row = [
                p["provider_name"], d.get("group", ""), d["label"], value,
                _STATUS_LABEL.get(f.get("status", ""), f.get("status", "")),
                f.get("confidence", ""),
                f.get("citation", ""), src.get("name", ""),
                prog.get("citation_url") if prog and not src else src.get("url", ""),
                _date(src.get("fetched_at", "")), (src.get("text_sha256", "") or "")[:12],
            ]
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = _WRAP_TOP
                cell.border = _BORDER
            r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    widths = [20, 20, 22, 60, 26, 11, 34, 30, 40, 12, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def write_segment_workbook(dataset: dict, group: str, title: str, path: str | Path,
                           view: str = "all") -> Path:
    """One-segment workbook (a single Comparison sheet) mirroring that section's
    on-site table, its per-segment dimension set, and the reader's active view."""
    wb = Workbook()
    _comparison_sheet(wb.active, dataset, group, title, view)
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return p


def write_workbook(dataset: dict, path: str | Path) -> Path:
    # One comparison sheet per segment, each with only its applicable dimensions,
    # so the workbook mirrors the site's per-segment tables. Plus a full Detail
    # sheet for provenance.
    wb = Workbook()
    first = True
    for group, title in (("cloud", "Cloud Infrastructure"),
                         ("closed", "Closed Model Providers"),
                         ("open_hosted", "Open - Hosted platforms"),
                         ("open_weights", "Open - Weights & licenses")):
        ws = wb.active if first else wb.create_sheet()
        first = False
        _comparison_sheet(ws, dataset, group, title)
    _detail_sheet(wb.create_sheet(), dataset)
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return p
